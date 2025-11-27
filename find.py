import time
import random
import os
import requests
import openpyxl
import urllib3
from selenium import webdriver
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading

# Tắt cảnh báo SSL
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# --- CẤU HÌNH ---
FOLDER_NAME = "hinh_anh_san_pham"
EXCEL_FILE = "DSSP.xlsx"
NUM_WORKERS = 8  # Số browser chạy song song (3 khuyến nghị, 4-5 nếu RAM >= 16GB)

# Lock để tránh xung đột khi ghi Excel
excel_lock = threading.Lock()

def setup_driver(worker_id=0):
    """Cấu hình Chrome Driver để tránh bị phát hiện là Bot"""
    chrome_options = Options()
    
    # Tạo thư mục profile riêng cho mỗi worker (tránh conflict)
    selenium_profile = os.path.join(os.getcwd(), f"selenium_profile_worker_{worker_id}")
    if not os.path.exists(selenium_profile):
        os.makedirs(selenium_profile)
    
    chrome_options.add_argument(f"user-data-dir={selenium_profile}")
    
    # Chạy ẩn (không hiện trình duyệt) - Bỏ comment dòng dưới nếu muốn chạy ngầm
    # chrome_options.add_argument("--headless") 
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--window-size=1920,1080")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    
    # Loại bỏ dấu hiệu Automation
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option('useAutomationExtension', False)
    
    # Đặt vị trí cửa sổ khác nhau cho mỗi worker (dễ theo dõi)
    chrome_options.add_argument(f"--window-position={worker_id * 100},{worker_id * 100}")

    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=chrome_options)
    return driver

def remove_vietnamese_accents(text):
    """Chuyển tiếng Việt có dấu thành không dấu"""
    vietnamese_map = {
        'à': 'a', 'á': 'a', 'ả': 'a', 'ã': 'a', 'ạ': 'a',
        'ă': 'a', 'ằ': 'a', 'ắ': 'a', 'ẳ': 'a', 'ẵ': 'a', 'ặ': 'a',
        'â': 'a', 'ầ': 'a', 'ấ': 'a', 'ẩ': 'a', 'ẫ': 'a', 'ậ': 'a',
        'è': 'e', 'é': 'e', 'ẻ': 'e', 'ẽ': 'e', 'ẹ': 'e',
        'ê': 'e', 'ề': 'e', 'ế': 'e', 'ể': 'e', 'ễ': 'e', 'ệ': 'e',
        'ì': 'i', 'í': 'i', 'ỉ': 'i', 'ĩ': 'i', 'ị': 'i',
        'ò': 'o', 'ó': 'o', 'ỏ': 'o', 'õ': 'o', 'ọ': 'o',
        'ô': 'o', 'ồ': 'o', 'ố': 'o', 'ổ': 'o', 'ỗ': 'o', 'ộ': 'o',
        'ơ': 'o', 'ờ': 'o', 'ớ': 'o', 'ở': 'o', 'ỡ': 'o', 'ợ': 'o',
        'ù': 'u', 'ú': 'u', 'ủ': 'u', 'ũ': 'u', 'ụ': 'u',
        'ư': 'u', 'ừ': 'u', 'ứ': 'u', 'ử': 'u', 'ữ': 'u', 'ự': 'u',
        'ỳ': 'y', 'ý': 'y', 'ỷ': 'y', 'ỹ': 'y', 'ỵ': 'y',
        'đ': 'd',
        'À': 'A', 'Á': 'A', 'Ả': 'A', 'Ã': 'A', 'Ạ': 'A',
        'Ă': 'A', 'Ằ': 'A', 'Ắ': 'A', 'Ẳ': 'A', 'Ẵ': 'A', 'Ặ': 'A',
        'Â': 'A', 'Ầ': 'A', 'Ấ': 'A', 'Ẩ': 'A', 'Ẫ': 'A', 'Ậ': 'A',
        'È': 'E', 'É': 'E', 'Ẻ': 'E', 'Ẽ': 'E', 'Ẹ': 'E',
        'Ê': 'E', 'Ề': 'E', 'Ế': 'E', 'Ể': 'E', 'Ễ': 'E', 'Ệ': 'E',
        'Ì': 'I', 'Í': 'I', 'Ỉ': 'I', 'Ĩ': 'I', 'Ị': 'I',
        'Ò': 'O', 'Ó': 'O', 'Ỏ': 'O', 'Õ': 'O', 'Ọ': 'O',
        'Ô': 'O', 'Ồ': 'O', 'Ố': 'O', 'Ổ': 'O', 'Ỗ': 'O', 'Ộ': 'O',
        'Ơ': 'O', 'Ờ': 'O', 'Ớ': 'O', 'Ở': 'O', 'Ỡ': 'O', 'Ợ': 'O',
        'Ù': 'U', 'Ú': 'U', 'Ủ': 'U', 'Ũ': 'U', 'Ụ': 'U',
        'Ư': 'U', 'Ừ': 'U', 'Ứ': 'U', 'Ử': 'U', 'Ữ': 'U', 'Ự': 'U',
        'Ỳ': 'Y', 'Ý': 'Y', 'Ỷ': 'Y', 'Ỹ': 'Y', 'Ỵ': 'Y',
        'Đ': 'D'
    }
    
    result = ""
    for char in text:
        result += vietnamese_map.get(char, char)
    return result

def download_image(url, filename, folder, image_number=1):
    """Hàm tải ảnh từ URL và lưu vào folder. Trả về đường dẫn đầy đủ đến file hoặc None nếu lỗi"""
    try:
        if not os.path.exists(folder):
            os.makedirs(folder)
            
        # Tắt SSL verification và thêm headers để tránh bị chặn
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        response = requests.get(url, stream=True, timeout=10, verify=False, headers=headers)
        if response.status_code == 200:
            # Bỏ dấu tiếng Việt
            no_accent = remove_vietnamese_accents(filename)
            # Chỉ giữ chữ, số và khoảng trắng
            clean_name = "".join([c for c in no_accent if c.isalnum() or c == ' ']).strip()
            # Thay khoảng trắng bằng dấu gạch dưới và thêm số thứ tự
            file_name = f"{clean_name.replace(' ', '_')}_{image_number}.jpg"
            file_path = os.path.join(folder, file_name)
            
            with open(file_path, 'wb') as f:
                f.write(response.content)
            print(f"[OK] Đã tải: {file_name}")
            
            # Trả về đường dẫn rút gọn: folder\filename.jpg
            relative_path = f"{folder}\\{file_name}"
            return relative_path
        else:
            print(f"[Lỗi] Không tải được ảnh: {filename}")
            return None
    except Exception as e:
        print(f"[Lỗi] Exception khi tải ảnh {filename}: {e}")
        return None

def read_products_from_excel(file_path):
    """Đọc danh sách sản phẩm từ Excel: cột 1 (barcode), cột 2 (name)"""
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        products = []
        # Đọc từ dòng 2 (bỏ qua header), đọc cả 2 cột
        for row in ws.iter_rows(min_row=2, min_col=1, max_col=2, values_only=True):
            barcode = str(row[0]).strip() if row[0] else None
            name = str(row[1]).strip() if row[1] else None
            
            if barcode and name:  # Cả 2 cột đều phải có giá trị
                products.append({
                    'barcode': barcode,
                    'name': name
                })
        
        wb.close()
        print(f">>> Đã đọc {len(products)} sản phẩm từ Excel")
        return products
    except Exception as e:
        print(f"[Lỗi] Không đọc được file Excel: {e}")
        return []

def write_image_paths_to_excel(file_path, row_index, image_paths):
    """Ghi đường dẫn các file ảnh vào cột 3, 4, 5 (img1, img2, img3) của Excel - Thread-safe"""
    with excel_lock:
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            # Ghi vào cột C, D, E (cột 3, 4, 5), dòng tương ứng
            for i, image_path in enumerate(image_paths, start=3):
                ws.cell(row=row_index, column=i, value=image_path)
            
            wb.save(file_path)
            print(f"[Worker] Đã ghi {len(image_paths)} đường dẫn vào Excel dòng {row_index}")
        except Exception as e:
            print(f"[Lỗi] Không ghi được vào Excel: {e}")

def process_product_thread(product_data, worker_id):
    """Xử lý một sản phẩm - hàm này sẽ được gọi bởi ThreadPoolExecutor"""
    product, index = product_data
    barcode = product['barcode']
    name = product['name']
    
    # Lấy thread ID để tạo profile riêng
    thread_id = threading.get_ident() % 1000
    profile_id = f"{worker_id}_{thread_id}"
    
    print(f"\n[Worker {worker_id}] --- Đang xử lý: {name} (Barcode: {barcode}) ---")
    
    driver = None
    try:
        # Delay ngẫu nhiên để tránh khởi động đồng thời quá nhiều Chrome instances
        startup_delay = random.uniform(0.3, 1.0)
        time.sleep(startup_delay)
        
        # Khởi động driver cho worker này
        driver = setup_driver(profile_id)
        
        # Đợi Chrome mở xong
        time.sleep(2)
        
        # Đóng các tab cũ nếu có
        if len(driver.window_handles) > 1:
            main_window = driver.window_handles[0]
            for handle in driver.window_handles[1:]:
                driver.switch_to.window(handle)
                driver.close()
            driver.switch_to.window(main_window)
        
        # Truy cập Google Images
        driver.get("https://www.google.com/imghp?hl=vi")
        time.sleep(2)
        
        # 1. Tạo URL tìm kiếm Google Images theo BARCODE + TÊN SẢN PHẨM
        search_query = f"{barcode} {name}"
        search_url = f"https://www.google.com/search?q={search_query.replace(' ', '+')}&tbm=isch&hl=vi"
        print(f"[Worker {worker_id}] >>> Tìm kiếm: {search_query}")
        driver.get(search_url)
        
        # 2. Random delay (giảm xuống vì có nhiều worker)
        delay = random.uniform(2, 3)
        print(f"[Worker {worker_id}] >>> Đợi {delay:.1f}s để trang load...")
        time.sleep(delay)
        
        # 3. Tìm tất cả ảnh có thể click được (tối đa 15 ảnh)
        print(f"[Worker {worker_id}] >>> Tìm ảnh trong kết quả...")
        
        # Thử nhiều selector khác nhau cho Google Images
        thumbnail_selectors = [
            '//img[@class="rg_i Q4LuWd"]',
            '//div[@jsname]//img',
            '//img[contains(@src, "gstatic")]',
            '//h3//ancestor::div[2]//img'
        ]
        
        all_thumbnails = []
        for selector in thumbnail_selectors:
            try:
                WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((By.XPATH, selector))
                )
                thumbnails = driver.find_elements(By.XPATH, selector)
                
                # Lấy tất cả ảnh có thể click (tối đa 15)
                for thumb in thumbnails[:15]:
                    try:
                        if thumb.is_displayed() and thumb.size['width'] > 50:
                            all_thumbnails.append(thumb)
                    except:
                        continue
                
                if len(all_thumbnails) >= 3:
                    break
            except:
                continue
        
        if not all_thumbnails:
            raise Exception("Không tìm thấy ảnh có thể click")
        
        print(f"[Worker {worker_id}] >>> Tìm được {len(all_thumbnails)} ảnh, cần lấy 3 ảnh")
        
        # 4. Lặp qua các ảnh và tải về cho đến khi đủ 3 ảnh
        downloaded_paths = []
        thumbnail_index = 0
        attempts = 0
        max_attempts = len(all_thumbnails)
        
        while len(downloaded_paths) < 3 and thumbnail_index < max_attempts:
            attempts += 1
            img_num = len(downloaded_paths) + 1
            thumbnail = all_thumbnails[thumbnail_index]
            
            try:
                print(f"[Worker {worker_id}] >>> Đang thử ảnh thứ {thumbnail_index + 1} (cần ảnh {img_num}/3)...")
                
                # Click vào ảnh bằng JavaScript
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", thumbnail)
                time.sleep(0.5)
                driver.execute_script("arguments[0].click();", thumbnail)
                time.sleep(2)
                
                # 5. Lấy ảnh full size từ panel bên phải
                all_images = driver.find_elements(By.TAG_NAME, "img")
                
                img_url = None
                max_size = 0
                
                for img in all_images:
                    try:
                        src = img.get_attribute("src")
                        if not src or "data:image" in src or "gstatic" in src:
                            continue
                        
                        # Lấy kích thước ảnh
                        width = img.size.get('width', 0)
                        height = img.size.get('height', 0)
                        size = width * height
                        
                        if size > max_size and "http" in src:
                            max_size = size
                            img_url = src
                    except:
                        continue
                
                # 6. Nếu không tìm thấy, thử lấy từ thumbnail
                if not img_url:
                    img_url = thumbnail.get_attribute("src")
                
                # 7. Download ảnh - Đặt tên theo NAME (cột 2) với số thứ tự
                if img_url and "http" in img_url:
                    image_path = download_image(img_url, name, FOLDER_NAME, img_num)
                    if image_path:
                        downloaded_paths.append(image_path)
                        print(f"[Worker {worker_id}] >>> ✓ Đã tải ảnh {img_num}/3")
                    else:
                        print(f"[Worker {worker_id}] >>> ✗ Không tải được, thử ảnh tiếp theo...")
                else:
                    print(f"[Worker {worker_id}] >>> ✗ Link không hợp lệ, thử ảnh tiếp theo...")
                
            except Exception as e:
                print(f"[Worker {worker_id}] >>> ✗ Lỗi: {str(e)[:50]}, thử ảnh tiếp theo...")
            
            # Chuyển sang ảnh tiếp theo
            thumbnail_index += 1
        
        # Kiểm tra kết quả
        if len(downloaded_paths) < 3:
            print(f"[Worker {worker_id}] >>> ⚠ Chỉ lấy được {len(downloaded_paths)}/3 ảnh sau {attempts} lần thử")
            # Điền các cột còn lại bằng thông báo
            while len(downloaded_paths) < 3:
                downloaded_paths.append("KHÔNG_TẢI_ĐƯỢC")
        else:
            print(f"[Worker {worker_id}] >>> ✓ Đã lấy đủ 3/3 ảnh sau {attempts} lần thử")
        
        # 8. Ghi tất cả đường dẫn vào Excel (cột 3, 4, 5)
        if downloaded_paths:
            write_image_paths_to_excel(EXCEL_FILE, index, downloaded_paths)
        else:
            write_image_paths_to_excel(EXCEL_FILE, index, ["KHÔNG TÌM THẤY", "", ""])
            # Lưu screenshot để debug
            screenshot_path = f"debug_{barcode[:20]}.png"
            driver.save_screenshot(screenshot_path)
        
        print(f"[Worker {worker_id}] ✓ Hoàn thành: {name}")
        return True
        
    except Exception as e:
        print(f"[Worker {worker_id}] [Lỗi] Có vấn đề khi xử lý {name}: {e}")
        write_image_paths_to_excel(EXCEL_FILE, index, ["LỖI", "LỖI", "LỖI"])
        # Lưu screenshot khi lỗi
        try:
            if driver:
                driver.save_screenshot(f"error_{barcode[:20]}.png")
        except:
            pass
        return False
    
    finally:
        # Đóng driver sau khi xử lý xong sản phẩm
        if driver:
            try:
                driver.quit()
            except:
                pass

def main():
    # Đọc danh sách sản phẩm từ Excel
    products = read_products_from_excel(EXCEL_FILE)
    
    if not products:
        print("[Lỗi] Không có sản phẩm nào để xử lý!")
        return
    
    print(f"\n{'='*60}")
    print(f">>> BẮT ĐẦU XỬ LÝ {len(products)} SẢN PHẨM VỚI {NUM_WORKERS} WORKERS")
    print(f"{'='*60}\n")
    
    # Chuẩn bị dữ liệu
    product_data = []
    for idx, product in enumerate(products, start=2):  # start=2 vì dòng 1 là header
        product_data.append((product, idx))
    
    # Sử dụng ThreadPoolExecutor thay vì multiprocessing
    # Threads chia sẻ memory nên Lock hoạt động tốt hơn trên Windows
    completed = 0
    total = len(product_data)
    
    with ThreadPoolExecutor(max_workers=NUM_WORKERS) as executor:
        # Submit tất cả tasks
        future_to_product = {
            executor.submit(process_product_thread, data, worker_id): data 
            for worker_id, data in enumerate(product_data)
        }
        
        # Xử lý kết quả khi hoàn thành
        for future in as_completed(future_to_product):
            completed += 1
            try:
                result = future.result()
                if result:
                    print(f"[Progress] {completed}/{total} sản phẩm hoàn thành")
            except Exception as e:
                print(f"[Lỗi] Task failed: {e}")
    
    print(f"\n{'='*60}")
    print(">>> HOÀN TẤT TẤT CẢ SẢN PHẨM!")
    print(f"{'='*60}\n")

if __name__ == "__main__":
    main()